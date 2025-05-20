import openpyxl
import pandas as pd
import argparse

def get_target_persons(filepath='data/standard_directory_persons.xlsx', sheetname='Person List', filter_red=False):
    # Load the workbook and sheet
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheetname]

    # Find the column index for 'Researcher/Date'
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = header.index('Researcher/Date') + 1  # openpyxl is 1-based

    # Collect rows (filter by red if requested)
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if filter_red:
            cell = row[col_idx - 1]
            fill = cell.fill
            # Check if the cell has a red fill (RGB 'FFFF0000' is standard red)
            if fill.fill_type is not None and fill.start_color.rgb in ('FFFF0000', 'FF0000'):
                rows.append([c.value for c in row])
        else:
            rows.append([c.value for c in row])

    # Convert to DataFrame
    df = pd.DataFrame(rows, columns=header)

    # Exclude records with data length >= 6 in any column starting with 'LOD' or 'lod'
    lod_cols = [col for col in df.columns if str(col).lower().startswith('lod')]
    mask = df[lod_cols].apply(lambda col: col.map(lambda x: len(str(x)) if pd.notnull(x) else 0))
    exclude_mask = (mask >= 6).any(axis=1)
    df_filtered = df[~exclude_mask]

    # If you want to not filter anything, just use the original DataFrame
    # df_filtered = df.copy()
    # Write to CSV in the data folder
    df_filtered.to_csv('data/target_persons.csv', index=False, encoding='utf-8-sig')

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--filter-red', action='store_true', help='Only include rows with red "Researcher/Date" cells')
    args = parser.parse_args()
    get_target_persons(filter_red=args.filter_red)
    print("target_persons.csv written to data folder.")